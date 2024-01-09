"""
Fitting script rewritten into OpenPNM V3.0.
Script fits the model current density to the experimental current density at a 
set of applied experimental (over)potentials (i.e. fitting of the polarization curve)

Simulates a mirrored cathode and anode electrode for an (extracted)  
network running isothermally at steady state. 
The dilute solution assumption is made, whereby the fluid transport is asssumed 
independent of the mass transport. The mass and charge transport are coupled
via the Butler-Volmer equation and are solved sequentially in an iterative manner
untill convergence of the current is achieved.  

Two flow fields can be simulated with this code including:
    1. The flow-through flow field: The FTFF flow field is modelled using boundary condition, and depletion over 
       the lenght of the network is modelled using a networks-in-series approach. Here
       the outlet concentration of the preceding network is imposed as the inlet 
       concentration of the current network.
    2. The interdigitated flow field: The IDFF flow field is modelled using boundary conditions. Depletion over the
       length of the network is neglected. Hence, the IDFF is modelled using a single 
       network. The modeling domain consists of:
           Half an inlet channel - a full rib - half an outlet channel. 
           At the inlet channel a flow rate is applied, and at the outlet channel
           a pressure boundary of 0 [Pa] is applied. At the rib a No flux Neumann
           condition is applied.
        
One can fit the following factors in this script: 
    1. km_factor            -> Mass transfer coefficient fitting factor. Note 
                               that for the  velocity dependent km correlation 
                               this represents a multiplication factor for the 
                               C1 factor of the km_{local} correlation.
    2. Conductivity_factor  -> Fitted conductivity fitting factor.
    Note that if fitting is *not* required set these fitting factors to unity (1).    
        
Working principle:
    1. Importing packages, set Load/Read directories, Loading in the Network, 
       and set/load simulation constants.
       
    2. Define and compute network properties, create phase objects with associated
       physics models and ascribe phase properties. 
       
    3. Load in stokesflow algorithms. These stokesflow algorithms were created in 
       the function "fluid field main", and their physics model is based on the
       mechanical energy balance in a pore-throat-pore element. 
       (Larachi et al. http://dx.doi.org/10.1016/j.cej.2013.11.077 )
                                
    4. Create and initialize mass and charge transport algorithms with 
       associated physics models and ascribe phase properties. 
    
    5. Fitting of the polarization curve data.
        a. Run mass and charge transport iterative algorithm sequentially in 
           each half-cell. The two half-cells are coupled via the membrane 
           potential.
           
        b. Checking convergence of the simulation. If converged compare simulation
           current density with the experimental current density input.
        
        c. Repeat step a-b up to finding the final fitting value(s) for the fitting
           parameters.
          
Input of script
    1. A .pnm network file
    2. An input dictionairy containing the parameters of the active species,
       phases, membrane, the applied voltage range, numeric convergence criteria,
       and possible a mass transfer correlation. 
    3. A set inlet velocity with associated current density
    4. Stokesflow data for a set inlet velocity
    5. A .CSV file containing experimental polarization curve data (current density - potential)
    
Options:  
    1. Larachi_HC
        If Larachi_HC == 0: OpenPNM physics are used (i.e. shape factor)
        If Larachi_HC == 1: Physics based on mechanical energy balance are used (Larachi et al.)
        *Note, however, that the OpenPNM physics pertain to the fluid field in 
        the altered network (where the throat diameter is set smaller or equal to diameter of connected pores).*
        
    2. Velocity_dependent_km
        If Velocity_dependent_km == 0: Velocity independent mass transfer 
            correlation km_{local} = R_{pore, i} / D_{active species}. Correlation 
            is adapted from Van der Heijden et al. DOI: 10.1149/1945-7111/ac5e46 
            
        If Velocity_dependent_km == 1: Velocity dependent mass transfer 
            correlation km_{local} = C1 * velocity_{pore}^{C2}. Here C1 and C2
            are the (fitted) prefactor and exponent, and velocity_{pore} the 
            velocity in the *pores*, which is derived following the methodology
            presented by Larachi et al. (http://dx.doi.org/10.1016/j.cej.2013.11.077 ).
            
    3. Flow field type
        If Flow_field == 0: The FTFF is simulated.
        If Flow_field == 1: The IDFF is simulated. 
"""

#---------------------------------Options-------------------------------------#
Larachi_HC = 1
Velocity_dependent_km = 1
Min_it_mass_charge = 20
Flow_field = 0

#------------------------------Fitting parameters-----------------------------#
# Note that the below fitting factors only pertain to initialization. 
km_factor =  1.0                
Rm_factor = 1.0
Conductivity_factor = 1.0

""" ---------Importing Packages and set Load/Read directories --------------"""
import openpnm as op
import numpy as np
import time
import scipy.optimize as optimize
import inputDictTEMPO as inputDict  
import openpyxl
import os
import Custom_functions_km as cf
import sys
import Custom_network as cn
import Custom_functions_transport as cf_trans
import Custom_functions_phase as cf_phase
import Custom_functions_pressure_fitting as cf_pres_fit  

# Setting import network directory
cwd = os.getcwd()                               # Curent Working directory
network_path = '\\Input_networks\\'             # Folder containing network
network_name = 'Freudenberg_check_CSV'          # Name of the network

# Setting file path directories for loading in networks
file_name_net_c = network_name + '_catholyte_network_throat_diameter_scaled.pnm'
file_name_net_a = network_name + '_anolyte_network_throat_diameter_scaled.pnm'
path_net_c = cwd + network_path + file_name_net_c
path_net_a = cwd + network_path + file_name_net_a
file_vec = [path_net_c]

# Setting exporting folder directory and simulation name
Output_folder = 'Folder'                        # Output folder 
name0_vec = ['Name']                            # Simulation name

# Setting simulation parameters
v_in_vec = [1.5e-2]                              # Inlet velocities [m/s]    
j0_vec = [1300]                                 # Current densities [A/m2]  
vel_ind = 0                                     # Velocity index for coupling exchange current density with a velocity

"""----------Loading in the Network and simulation constants----------------"""
for NrNetwork, file in enumerate(file_vec):
    for v_in in v_in_vec:
        j0 = j0_vec[vel_ind]    
        name = name0_vec[NrNetwork] + '_' + str(v_in*100) + '_cm_s-1' + '_j0_' + str(j0)
        
        # Create output directory for saving the .VTK files. These are created 
        # within the folder ".\\output\\Output_folder" where Output_folder is set above.
        # os.mkdir('.\\output\\' + Output_folder + '\\' + name)  
        starting_time = time.time()
        op.Workspace().clear()  
        
       # Loading the cathodic and anodic network
        file_name = file + '.pnm'
        project_cat = op.Workspace().load_project(filename = path_net_c)
        net_c = project_cat.network   
        project_ano = op.Workspace().load_project(filename = path_net_a)
        net_a = project_ano.network                
    
        # Checking for isolated and disconnected pores:
        health_c = op.utils.check_network_health(network = net_c)
        print("The health of the cathodic and anodic electrodes are as",
              "follows (note that they are the same mirrored network):\n")
        print(health_c); 

        # Update the face labels that are defined in the SNOW algorithm
        cf.validate_face_labels(net_c) 
        cf.validate_face_labels(net_a)

        # Instantiate the output Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        polarizationCurveData = {}
        
        # Load in the input dictionary
        param = inputDict.input_dict 

        """-----------Define and compute network properties-------------"""
        # Define electrode dimensions [x=0, y=1, z=2]
        # NOTE: Make sure you know which dimension correlates with which face in your image.
        H_dim = param['height_dimension']       # The dimension corresponding to the height of the electrodes (sides of the electrode)
        L_dim = param['length_dimension']       # The dimension corresponding to the length of the electrodes (flow field direction)
        W_dim = param['width_dimension']        # The dimension corresponding to the width of the electrodes (thickness: current collector -> membrane)            
        
        L_c = np.amax(net_c['pore.coords'][net_c.pores(), L_dim])       # Maximum Length [m]
        H_c = np.amax(net_c['pore.coords'][net_c.pores(), H_dim])       # Maximum Height [m]
        H_c_min = np.amin(net_c['pore.coords'][net_c.pores(), H_dim])   # Minimum Height [m] (offset from y = 0)
        W_c = np.amax(net_c['pore.coords'][net_c.pores(), W_dim])       # Maximum Width [m]
        L_a = np.amax(net_a['pore.coords'][net_a.pores(), L_dim])
        H_a = np.amax(net_a['pore.coords'][net_a.pores(), H_dim])
        H_a_min = np.amin(net_a['pore.coords'][net_a.pores(), H_dim])
        W_a = np.amax(net_a['pore.coords'][net_a.pores(), W_dim])
        
        A_ext_c = L_c * H_c                     # Cathode area to compute external current density [m2]
        A_ext_a = L_a * H_a                     # Anode area [m2]
        mem_area = A_ext_c                      # Fictitious membrane area [m2]  
        
        if Flow_field == 0:
            A_in_c = W_c * H_c                  # Cathode inlet area [m2]
            A_in_a = W_a * H_a                  # Anode inlet area [m2]
        elif Flow_field == 1:
            A_in_c = W_c * L_c                  # Cathode inlet area [m2]
            A_in_a = W_a * L_a                  # Anode inlet area [m2]            
        
        # Compute electrochemically active surface area 
        Ai_c = net_c['pore.surface_area']       # Cathode internal surface area [m2]
        Ai_a = net_a['pore.surface_area'] 
        
        # Pore radii
        rp_c = net_c['pore.diameter'] / 2       # Pore radii in the cathode [m]
        rp_a = net_a['pore.diameter'] / 2       # Pore radii in the anode [m]            
        
        """-----------Define and compute phase properties---------------"""
        #-------------------Adding phase constants------------------------#
        # Phase constants are assigned via models, as during phase
        # regeneration the water object inherent functions would overwrite 
        # assigned phase constants.
        
        Conductivity_value = param['anolyte_conductivity'] * Conductivity_factor

        anolyte = op.phase.Water(network=net_a, name='anolyte')
        anolyte.add_model(propname = 'pore.electrical_conductivity',                # Anolyte electrical conductivity [S m-1]
                  model = cf_phase.custom_electrical_conductivity_fit,
                  conductivity_fitted = Conductivity_value)                  
        anolyte.add_model(propname = 'pore.viscosity',                              # Anolyte viscosity [Pa s]   
                          model = cf_phase.custom_viscosity,
                          parameter_script = param, prop = 'anolyte_viscosity')
        anolyte.add_model(propname = 'pore.density',                                # Anolyte density [kg m-3]  
                          model = cf_phase.custom_density,
                          parameter_script = param, prop = 'anolyte_density')
        anolyte.add_model(propname = 'pore.diffusivity',                            # Anolyte active species diffusivity in electrolyte [m2 s-1]
                          model = cf_phase.custom_diffusivity,
                          parameter_script = param, prop = 'D_a')            
        anolyte['pore.molecular_weight'] = 0.01802

        catholyte = op.phase.Water(network=net_c, name='catholyte')
        catholyte.add_model(propname = 'pore.electrical_conductivity',              # Catholyte electrical conductivity [S m-1]
                          model = cf_phase.custom_electrical_conductivity_fit,
                          conductivity_fitted = Conductivity_value)  
        catholyte.add_model(propname = 'pore.diffusivity',                          # Catholyte active species diffusivity in electrolyte [m2 s-1]
                          model = cf_phase.custom_diffusivity,
                          parameter_script = param, prop = 'D_c')           
        catholyte.add_model(propname = 'pore.viscosity',                            # Catholyte viscosity [Pa s]
                          model = cf_phase.custom_viscosity,
                          parameter_script = param, prop = 'catholyte_viscosity')                 
        catholyte.add_model(propname = 'pore.density',                              # Catholyte density [kg m-3]  
                          model = cf_phase.custom_density,
                          parameter_script = param, prop = 'catholyte_density')
        catholyte['pore.molecular_weight'] = 0.01802
        
        """-----------------------Stokes algorithm----------------------"""
        # Loading the pre-run stokes algorithm. Note that some phase models require
        # certain phase properties (i.e advection-diffusion required hydraulic conductance). 
        # Hence, the SF-algorithm should be initialized prior to assigning these models.
        
        # Creating stokes flow algorithm
        sf_c_odd = op.algorithms.StokesFlow(network=net_c, phase=catholyte)
        sf_c_even = op.algorithms.StokesFlow(network=net_c, phase=catholyte)
        sf_a_odd = op.algorithms.StokesFlow(network=net_a, phase=anolyte)
        sf_a_even = op.algorithms.StokesFlow(network=net_a, phase=anolyte)
        
        # Setting input paths
        Folder_SF = '\\Input_Fluid_field\\'
        file_name_sf_c_odd = 'SF_' + str(v_in*100).replace(".", "_") + 'cms_catholyte_odd_network'
        file_name_sf_a_odd = 'SF_' + str(v_in*100).replace(".", "_") + 'cms_anolyte_odd_network'
        file_name_sf_c_even = 'SF_' + str(v_in*100).replace(".", "_") + 'cms_catholyte_even_network'
        file_name_sf_a_even = 'SF_' + str(v_in*100).replace(".", "_") + 'cms_anolyte_even_network'
        
        path_sf_c_odd = cwd + Folder_SF + file_name_sf_c_odd + '.pnm'
        path_sf_a_odd = cwd + Folder_SF + file_name_sf_a_odd + '.pnm'
        path_sf_c_even = cwd + Folder_SF + file_name_sf_c_even + '.pnm'
        path_sf_a_even = cwd + Folder_SF + file_name_sf_a_even + '.pnm'
        
        # Importing stokes flow simulation data
        cf_pres_fit.Import_stokesflow_algo(sf_c_odd, 1, path_sf_c_odd)
        cf_pres_fit.Import_stokesflow_algo(sf_a_odd, 1, path_sf_a_odd)
        cf_pres_fit.Import_stokesflow_algo(sf_c_even, 0, path_sf_c_even)
        cf_pres_fit.Import_stokesflow_algo(sf_a_even, 0, path_sf_a_even)
        
        # Updating phase with sf_c_odd
        if Larachi_HC == 0:         # OpenPNM physics
            cf_pres_fit.Update_phase_with_SF(sf_c_odd, catholyte, 1, 0)
            cf_pres_fit.Update_phase_with_SF(sf_a_odd, anolyte, 1, 0)        
            
            # Compute and update phase and associated algorithms with mass transfer coefficient
            cf_pres_fit.Mass_transfer_coefficient(net_c, catholyte, 1, Velocity_dependent_km, 0, param) 
            cf_pres_fit.Mass_transfer_coefficient(net_a, anolyte, 1, Velocity_dependent_km, 0, param)
            
        elif Larachi_HC == 1:       # Mechanical energy balance (Larachi et al.)
            cf_pres_fit.Update_phase_with_SF(sf_c_odd, catholyte, 1, 1)
            cf_pres_fit.Update_phase_with_SF(sf_a_odd, anolyte, 1, 1)    
            
            # Compute and update phase and associated algorithms with mass transfer coefficient
            cf_pres_fit.Mass_transfer_coefficient(net_c, catholyte, 1, Velocity_dependent_km, 1, param)
            cf_pres_fit.Mass_transfer_coefficient(net_a, anolyte, 1, Velocity_dependent_km, 1, param)
        
        #---------------------Adding phase physics------------------------#
        # Custom functions (copy of those used in OpenPNM V2.2.0) are used
        # due to discrepencies w.r.t. the diameter of the throat exceeding
        # that of the pores from the network extraction.
        
        f_poisson = cf_trans.Poisson_shape_factors_ball_and_stick
        anolyte.add_model(propname = 'throat.poisson_shape_factors', model = f_poisson)
        catholyte.add_model(propname = 'throat.poisson_shape_factors', model = f_poisson)
        f_diff_cond = cf_trans.Diffusive_conductance_mixed_diffusion
        anolyte.add_model(propname = 'throat.diffusive_conductance', model = f_diff_cond)
        catholyte.add_model(propname = 'throat.diffusive_conductance', model = f_diff_cond)
        f_ad_dif = cf_trans.Advection_diffusion
        anolyte.add_model(propname = 'throat.ad_dif_conductance', model = f_ad_dif)
        catholyte.add_model(propname = 'throat.ad_dif_conductance', model = f_ad_dif)
        f_elec_con = cf_trans.Electrical_conductance_series_resistors
        anolyte.add_model(propname = 'throat.electrical_conductance', model = f_elec_con)            
        catholyte.add_model(propname = 'throat.electrical_conductance', model = f_elec_con)            
        
        anolyte.regenerate_models()
        catholyte.regenerate_models()            
        
        """-------Load in parameters for mass and charge transport------"""
        # Active species / kinetic parameters
        conc_in_a0 = param['conc_in_a']         # Anolyte active species inlet concentration in network 0 [mol m-3]
        conc_in_c0 = param['conc_in_c']         # Catholyte active species inlet concentration in network 0 [mol m-3]
        
        # Cell potential parameters (E_cell < 0 --> Charging, E_cell > 0 --> Discharging)
        E_red_a = param['E_red_a']              # Standard reduction potential of the anodic half reaction [V]
        E_red_c = param['E_red_c']              # Standard reduction potential of the cathodic half reaction [V]
        V_step = param['V_step']                # Step change in the cell voltage [V] 
        E_cell_final = param['E_cell_final']    # Final value of the cell voltage range [V] 
        
        # Potential parameters
        E_0 = E_red_c - E_red_a                                             # Open circuit voltage [V]
        E_cell_vec = np.arange(E_0 + V_step, E_cell_final+V_step, V_step)   # Cell voltage range for which the algorithm calculates
    
        # Membrane ionic membrane resistivity [Ohm m2] defined through Pouillet's law of electricity
        Conductivity_value = catholyte['pore.electrical_conductivity'][0]
        R_mem = cf.Membrane_conductivity(Conductivity_value, param)
        res_mem = R_mem / mem_area              # Membrane resistance [Ohm]
                   
        """--------------Mass and charge transport models---------------"""
        # Required networks in series for encompassing the total electrode length.
        # Number of networks
        if Flow_field == 0:
            number_of_networks = int(np.ceil(param['total_electrode_length']/L_c)) 
        if Flow_field == 1:
            number_of_networks = 1
        
        if Flow_field == 0:
            # Set-up advection-Diffusion algorithm
            ad_c_odd = op.algorithms.AdvectionDiffusion(network=net_c, phase=catholyte)
            ad_c_even = op.algorithms.AdvectionDiffusion(network=net_c, phase=catholyte)
            ad_a_odd = op.algorithms.AdvectionDiffusion(network=net_a, phase=anolyte)
            ad_a_even = op.algorithms.AdvectionDiffusion(network=net_a, phase=anolyte)            
            
            # Setting boundary conditions: No Flux B.C. between the networks 
            # These are updated in the iterative loop: the inlet boundary 
            # conditions of the even networks (2, 4) are the outlet boundary 
            # conditions of the odd networks (1, 3).
            ad_c_odd.set_outflow_BC(pores=net_c.pores('flow_outlet'))
            ad_c_even.set_outflow_BC(pores=net_c.pores('flow_inlet'))
            ad_a_odd.set_outflow_BC(pores=net_a.pores('flow_outlet'))
            ad_a_even.set_outflow_BC(pores=net_a.pores('flow_inlet'))            
        
        if Flow_field == 1:
            # Set-up advection-Diffusion algorithm
            ad_c_odd = op.algorithms.AdvectionDiffusion(network=net_c, phase=catholyte)
            ad_a_odd = op.algorithms.AdvectionDiffusion(network=net_a, phase=anolyte)
            if number_of_networks > 1:  # In case there are more than 1 network (Avoid Nans outputted in .vtk file)
                ad_c_even = op.algorithms.AdvectionDiffusion(network=net_c, phase=catholyte)
                ad_a_even = op.algorithms.AdvectionDiffusion(network=net_a, phase=anolyte)
                
            # Setting boundary conditions: No Flux B.C. between the networks 
            # These are updated in the iterative loop: the inlet boundary 
            # conditions of the even networks (2, 4) are the outlet boundary 
            # conditions of the odd networks (1, 3).
            ad_c_odd.set_outflow_BC(pores=net_c.pores('flow_outlet'))
            ad_a_odd.set_outflow_BC(pores=net_a.pores('flow_outlet'))
            if number_of_networks > 1:  # In case there are more than 1 network (Avoid Nans outputted in .vtk file)
                ad_c_even.set_outflow_BC(pores=net_c.pores('flow_outlet'))
                ad_a_even.set_outflow_BC(pores=net_a.pores('flow_outlet')) 
          
        # Source term: Butler-Volmer equation for the concentration. The 
        # linear source term is defined as: r = A_{1} X + A_{2}. The inter-
        # cept  A_{2} = 0. A_{1} will be defined in the iterative loop via 
        # the function cf.cf.bv_rate_constant_ad_c/a and is the non-linearized
        # BV equation including mass transfer. The independent variable
        # X is the concentration inside the pores.
        source_term = op.models.physics.source_terms.linear
        anolyte['pore.rxn_A2'] = 0.0
        catholyte['pore.rxn_A2'] = 0.0
        anolyte.add_model(propname='pore.butler_volmer', 
                          model=source_term,
                          A1='pore.rxn_A1', A2="pore.rxn_A2",
                          X='pore.concentration',
                          regen_mode="deferred")
        catholyte.add_model(propname='pore.butler_volmer', 
                            model=source_term,
                            A1='pore.rxn_A1', 
                            A2="pore.rxn_A2", 
                            X='pore.concentration',
                            regen_mode="deferred")            
                
        if Flow_field == 0:
            ad_c_odd.set_source(propname='pore.butler_volmer', pores=net_c.pores('internal'))
            ad_c_even.set_source(propname='pore.butler_volmer', pores=net_c.pores('internal'))
            ad_a_odd.set_source(propname='pore.butler_volmer', pores=net_a.pores('internal'))
            ad_a_even.set_source(propname='pore.butler_volmer', pores=net_a.pores('internal'))            
        
        if Flow_field == 1:
            ad_c_odd.set_source(propname='pore.butler_volmer', pores=net_c.pores('internal'))
            ad_a_odd.set_source(propname='pore.butler_volmer', pores=net_a.pores('internal'))
            if number_of_networks > 1:  # In case there are more than 1 network (Avoid Nans outputted in .vtk file)
                ad_c_even.set_source(propname='pore.butler_volmer', pores=net_c.pores('internal'))
                ad_a_even.set_source(propname='pore.butler_volmer', pores=net_a.pores('internal'))          
        
        # Set-up ohmic conduction algorithm
        oc_c = op.algorithms.OhmicConduction(network=net_c, phase=catholyte)
        oc_a = op.algorithms.OhmicConduction(network=net_a, phase=anolyte)            
        # A note on membrane boundary conditions: Boundary condition at membrane 
        # is a function of the current and can therefore be found in the iteration 
        # loop, other boundaries are automatically set to no flux, i.e. dV/dx = 0.
        
        # Source term: Butler-Volmer equation for the potential. Use is 
        # made of the *linearized* BV equation including mass transfer. The
        # equation is: S = dIdv * V_new_a + I_ad_dif - dIdv * V_old_a. Here
        # A1 = dIdv and A2 ( = constant) = I_ad_dif - dIdv * V_old_a.
        # The independent variable X is the concentration inside the pores. 
        anolyte.add_model(propname='pore.proton_anolyte', 
                          model=source_term,
                          A1='pore.current_anolyte_A1',
                          A2='pore.current_anolyte_A2',
                          X='pore.voltage',
                          regen_mode="deferred")
        
        catholyte.add_model(propname='pore.proton_catholyte', 
                            model=source_term,
                            A1='pore.current_catholyte_A1',
                            A2='pore.current_catholyte_A2',
                            X='pore.voltage',
                            regen_mode="deferred")
        
        oc_c.set_source(propname='pore.proton_catholyte', pores=net_c.pores('internal'))
        oc_a.set_source(propname='pore.proton_anolyte', pores=net_a.pores('internal'))                 
        
        """------------Initializing mass and charge transport-----------"""

        def iterative_solver(params, anolyte, catholyte, net_c, net_a, E_cell_vec, rp_a, rp_c, j0, data, oc_c, oc_a,
                                                                                ad_c_even, ad_c_odd, ad_a_even, ad_a_odd, sf_c_odd, sf_c_even, sf_a_odd, sf_a_even, param, Velocity_dependent_km, Min_it_mass_charge, Larachi_HC):
            
            # Ascribing fitting values. Note that if you do *not* want to fit
            # a certain factor set this factor to unity.
            km_factor = 1.0     # params['kmfactor'].value # For vel-dep km: fit for pre-exponential factor of *Local* MT correlation     
            Conductivity_factor = params['Conductivityfactor'].value    # Conductivity fitting factor
            Rm_factor = 1.0     # params['Rmfactor'].value # Membrane resistance factor
            SA_factor = 1.0     # params['SAfactor'].value # Relative Surface area (SA) factor
            
            print('xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx\nThe Conductivity_factor is:', Conductivity_factor, '\nxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx')

            # Re-assign fitting conductivity to anolyte and catholyte
            Conductivity_fitted_value = param['anolyte_conductivity'] * Conductivity_factor
            anolyte.add_model(propname = 'pore.electrical_conductivity',                # Anolyte electrical conductivity [S m-1]
                              model = cf_phase.custom_electrical_conductivity_fit,
                              conductivity_fitted = Conductivity_fitted_value)    
            
            catholyte.add_model(propname = 'pore.electrical_conductivity',              # Catholyte electrical conductivity [S m-1]
                              model = cf_phase.custom_electrical_conductivity_fit,
                              conductivity_fitted = Conductivity_fitted_value) 
            anolyte.regenerate_models()
            catholyte.regenerate_models()    
            
            # Membrane ionic membrane resistivity [Ohm m2] defined through Pouillet's law of electricity
            Electrolyte_conductivity = catholyte['pore.electrical_conductivity'][0]
            R_mem = cf.Membrane_conductivity(Electrolyte_conductivity, param)
            res_mem = R_mem / mem_area                      # Membrane resistance [Ohm]
            
            # Compute electrochemically active surface area (=Geometric internal surface area)
            Ai_c = net_c['pore.surface_area'] * SA_factor   # Cathode internal surface area [m2]
            Ai_a = net_a['pore.surface_area'] * SA_factor
            
            # Initialize vectors for saving the current density. Used for the 
            # fitting of the model to the experimental data.
            current_sum_vec = np.zeros([1, len(E_cell_vec)])
            current_sum_vec_int = np.zeros([1, number_of_networks])    
            
            # Initialize total domain overpotential as zero and concentration
            # as inlet concentration. In further iterations, initial guesses 
            # will be converged values of previous cell voltage.
            eta0_a = np.zeros((net_a.Np, number_of_networks))
            eta0_c = np.zeros((net_c.Np, number_of_networks))            
            conc0_a = np.ones((net_a.Np, number_of_networks)) * conc_in_a0 
            conc0_c = np.ones((net_c.Np, number_of_networks)) * conc_in_c0            
            
            # Matrixes to store inlet concentration for the odd and even networks
            if Flow_field == 0:
                conc0_in_a_odd = np.zeros((len(net_a.pores('flow_inlet')), number_of_networks))
                conc0_in_c_odd = np.zeros((len(net_c.pores('flow_inlet')), number_of_networks))
                conc0_in_a_even = np.zeros((len(net_a.pores('flow_outlet')), number_of_networks))
                conc0_in_c_even = np.zeros((len(net_c.pores('flow_outlet')), number_of_networks))            
            if Flow_field == 1:
                conc0_in_a_odd = np.zeros((len(net_a.pores('flow_inlet')), number_of_networks))
                conc0_in_c_odd = np.zeros((len(net_c.pores('flow_inlet')), number_of_networks))
                conc0_in_a_even = np.zeros((len(net_a.pores('flow_inlet')), number_of_networks))
                conc0_in_c_even = np.zeros((len(net_c.pores('flow_inlet')), number_of_networks))            
                
            # Set the inlet concentration of the 0th network to the inlet concentration of the electrode
            conc0_in_a_odd[:, 0] = conc_in_a0
            conc0_in_c_odd[:, 0] = conc_in_c0            
            
            rel_tol = param['rel_tol']                  # Relative tolerance
            abs_tol = param['abs_tol']                  # Absolute tolerance for external current density
            max_iter = param['max_iter']                # Maximum number of iterations
            omega = param['omega']                      # Damping factor            
            
            """---Start of mass and charge transport iterative algorithm----"""
            # In this iterative algorithm we loop over each assigned cell 
            # voltage and network in series and solve iteratively the coupled
            # mass and charge balances.
            
            for E_cell_idx, E_cell in enumerate(E_cell_vec):
                for network in range(number_of_networks):
                    
                    # Update guesses for the overpotential and concentration
                    # based on the previous result
                    eta_c = eta0_c[:, network]          # Overpotential [V]
                    eta_a = eta0_a[:, network]
                    
                    # Assigning inlet concentrations. Note: network+1 for the
                    # network number since the range function starts counting at network 0
                    if (network + 1) % 2 == 0:          # Even network
                        conc_in_a = conc0_in_a_even[:, network] 
                        conc_in_c = conc0_in_c_even[:, network]
                    else:                               # Odd network
                        conc_in_a = conc0_in_a_odd[:, network]
                        conc_in_c = conc0_in_c_odd[:, network]
                        
                    # Initialize concentration in all pores based on previous result [mol m-3]
                    conc_c = conc0_c[:, network]  
                    conc_a = conc0_a[:, network]
                    
                    # Resetting current (to reset the relative error)
                    current_ad_a = 0.0                  # Current [A]
                    current_oc_a = 0.0
                    current_ad_c = 0.0
                    current_oc_c = 0.0                    
                    
                    # Initializing potential
                    V_c = E_cell - E_0 - eta_c          # Voltage in the liquid phase [V]
                    catholyte['pore.voltage'] = V_c 
                    V_a = 0 - eta_a                    
                    anolyte['pore.voltage'] = V_a
                    
                    # Flip the network for visualization in paraview
                    inversion_factor = net_a['pore.coords'][:, L_dim] - min(net_a['pore.coords'][:, L_dim])
                    new_position = max(net_a['pore.coords'][:, L_dim]) - inversion_factor
                    net_a['pore.coords'][:, L_dim] = new_position
                    net_c['pore.coords'][:, L_dim] = new_position                    
                    
                    # Start of iterative loop. First we update the cathodic 
                    # half-cell, then the anodic half-cell. 
                    for itr in range(max_iter):
                    
                    #------------------ Cathodic half-cell -----------------------#  
                        # Solve the advection-diffusion-reaction equation for the cathode
                        if (network + 1) % 2 == 0:      # Even networks
                            # Retrieve the StokesFlow data for an even network
                            catholyte.update(sf_c_even.soln)
                            
                            if Larachi_HC == 0:         # OpenPNM fluid field physics
                                cf_pres_fit.Update_phase_with_SF(sf_c_even, catholyte, 0, 0)
                                # MT coeficients                            
                                cf_pres_fit.Mass_transfer_coefficient(net_c, catholyte, 0, Velocity_dependent_km, 0, param)
                                
                            elif Larachi_HC == 1:       # Mechanical energy balance fluid field
                                cf_pres_fit.Update_phase_with_SF(sf_c_even, catholyte, 0, 1) 
                                # MT coeficients                            
                                cf_pres_fit.Mass_transfer_coefficient(net_c, catholyte, 0, Velocity_dependent_km, 1, param)
                                 
                            # Regenerate (the ad_dif_conductance) models
                            catholyte.regenerate_models('throat.ad_dif_conductance')
                            
                            # Set the boundary conditions for even networks (reverted case)
                            if Flow_field == 0:
                                ad_c_even.set_value_BC(pores = net_c.pores('internal', mode = 'nor'), mode='remove')
                                ad_c_even.set_value_BC(values=conc_in_c, pores=net_c.pores('flow_outlet'), mode='overwrite')
                                ad_c_even.set_outflow_BC(pores = net_c.pores('internal', mode = 'nor'), mode='remove')
                                ad_c_even.set_outflow_BC(pores=net_c.pores('flow_inlet'), mode='overwrite')
                            if Flow_field == 1:
                                ad_c_even.set_value_BC(pores = net_c.pores('internal', mode = 'nor'), mode='remove')
                                ad_c_even.set_value_BC(values=conc_in_c, pores=net_c.pores('flow_inlet'), mode='overwrite')
                                ad_c_even.set_outflow_BC(pores = net_c.pores('internal', mode = 'nor'), mode='remove')
                                ad_c_even.set_outflow_BC(pores=net_c.pores('flow_outlet'), mode='overwrite')
                            
                            # Update ad reaction term with new values for eta_c
                            catholyte['pore.rxn_A1'] = cf.bv_rate_constant_ad_c(eta_c, Ai_c, rp_c, j0, km_factor, catholyte['pore.km_exp'], param)
                            
                            # Solve the advection-diffusion-reaction equation
                            if itr > 0:
                                conc_guess = ad_c_even['pore.concentration']
                                ad_c_even.run(x0 = conc_guess)
                            else:
                                ad_c_even.run()
                                
                            # Update the concentration
                            conc_new_c = ad_c_even['pore.concentration']  
                            
                        else:                           # Odd networks
                            # Retrieve the StokesFlow data for an odd network
                            catholyte.update(sf_c_odd.soln)
                            
                            if Larachi_HC == 0:         # OpenPNM fluid field physics
                                cf_pres_fit.Update_phase_with_SF(sf_c_odd, catholyte, 1, 0)
                                # MT coeficients                            
                                cf_pres_fit.Mass_transfer_coefficient(net_c, catholyte, 1, Velocity_dependent_km, 0, param)
                                
                            elif Larachi_HC == 1:       # Mechanical energy balance fluid field
                                cf_pres_fit.Update_phase_with_SF(sf_c_odd, catholyte, 1, 1) 
                                # MT coeficients                            
                                cf_pres_fit.Mass_transfer_coefficient(net_c, catholyte, 1, Velocity_dependent_km, 1, param)
    
                            # Regenerate the ad_dif_conductance model
                            catholyte.regenerate_models('throat.ad_dif_conductance')
                            
                            # Set the boundary conditions for odd networks (normal case)
                            ad_c_odd.set_value_BC(pores = net_c.pores('internal', mode = 'nor'), mode='remove')
                            ad_c_odd.set_value_BC(values=conc_in_c, pores=net_c.pores('flow_inlet'), mode='overwrite')
                            ad_c_odd.set_outflow_BC(pores = net_c.pores('internal', mode = 'nor'), mode='remove')
                            ad_c_odd.set_outflow_BC(pores=net_c.pores('flow_outlet'), mode='overwrite')
                            
                            # Update ad reaction term with new values for eta_c
                            catholyte['pore.rxn_A1'] = cf.bv_rate_constant_ad_c(eta_c, Ai_c, rp_c, j0, km_factor, catholyte['pore.km_exp'], param)
                            
                            # Solve the advection-diffusion-reaction equation
                            if itr > 0:
                                conc_guess = ad_c_odd['pore.concentration']
                                ad_c_odd.run(x0 = conc_guess)
                            else:
                                ad_c_odd.run()      
                                
                            # Update the concentration
                            conc_new_c = ad_c_odd['pore.concentration']
                        
                        # Update the value for the concentration
                        conc_c = conc_new_c * omega + conc_c * (1 - omega)
                        catholyte['pore.concentration'] = conc_c                        
                        
                        # Compute TOTAL half-cell current according to the reaction of species
                        # NOTE: we only use the internal pores for the relative error calculation, 
                        #       as both the ohmic conduction and advection-diffusion algorithm use different boundary pores.
                        current_estimation_ad_c = cf.bv_rate_constant_oc_c(conc_c, eta_c, Ai_c, rp_c, j0, km_factor, catholyte['pore.km_exp'], param)[net_c.pores('internal')].sum()                      
                        
                        # Compute oc reaction term with new values for conc_c
                        drdv = cf.bv_rate_derivative_oc_c(conc_c, eta_c, Ai_c, rp_c, j0, km_factor, catholyte['pore.km_exp'], param)
                        r = cf.bv_rate_constant_oc_c(conc_c, eta_c, Ai_c, rp_c, j0, km_factor, catholyte['pore.km_exp'], param)
                        
                        catholyte['pore.current_catholyte_A1'] = drdv
                        catholyte['pore.current_catholyte_A2'] = r - drdv * V_c   
                        
                        # Couple the potential field in both half cells via a 
                        # continuum averaged membrane approach, in which we use 
                        # the pore potential at the anodic side to compute the 
                        # pore potential at the cathodic side. 
                        # This only works when identical (mirrored) anodic and cathodic pore.ele
                        # networks are used (networks with identical membrane pores)
                        # (the membrane pores remain on the membrane side for both odd and even networks)
                        V_membrane_pores_a = anolyte['pore.voltage'][net_a.pores('membrane')]                    # List of potential in the anodic pores next to the membrane 
                        V_membrane_pores_c = V_membrane_pores_a + res_mem * Rm_factor * current_estimation_ad_c  # List of potential in the cathodic pores next to the membrane                  
                        
                        # Update the membrane boundary condition for the cathode
                        oc_c.set_value_BC(values=V_membrane_pores_c, pores=net_c.pores('membrane'), mode='overwrite')                        
                        
                        # Solve the ohmic conduction equation (potential field) for the liquid electrolyte in the porous cathode
                        oc_c.run()
            
                        # Update the value for the liquid potential
                        V_new_c = oc_c['pore.voltage']
                        V_c = V_new_c * omega + V_c * (1 - omega)
                        catholyte['pore.voltage'] = V_c
                        
                        # Compute current according to the transport of species
                        current_estimation_oc_c = cf.bv_rate_constant_oc_c(conc_c, eta_c, Ai_c, rp_c, j0, km_factor, catholyte['pore.km_exp'], param)[net_c.pores('internal')].sum()
            
                        # Compute new cathodic overpotential
                        eta_c = E_cell - V_c - E_0                        
                        
                      #------------------ Anodic half-cell -----------------------#       
                        # Solve the advection-diffusion-reaction equation for the anode
                        if (network + 1) % 2 == 0:
                            # Retrieve the StokesFlow data for an even network
                            anolyte.update(sf_a_even.soln)
                            
                            if Larachi_HC == 0:             # OpenPNM fluid field physics
                                cf_pres_fit.Update_phase_with_SF(sf_a_even, anolyte, 0, 0)
                                # MT coeficients                            
                                cf_pres_fit.Mass_transfer_coefficient(net_a, anolyte, 0, Velocity_dependent_km, 0, param)         
                                
                            elif Larachi_HC == 1:           # Mechanical energy balance fluid field
                                cf_pres_fit.Update_phase_with_SF(sf_a_even, anolyte, 0, 1) 
                                # MT coeficients                            
                                cf_pres_fit.Mass_transfer_coefficient(net_a, anolyte, 0, Velocity_dependent_km, 1, param)
                                
                            # Regenerate the (ad_dif_conductance) models
                            anolyte.regenerate_models('throat.ad_dif_conductance')
                            
                            # Set the boundary conditions for even networks (reverted case)
                            if Flow_field == 0:
                                ad_a_even.set_value_BC(pores = net_a.pores('internal', mode = 'nor'), mode='remove')
                                ad_a_even.set_value_BC(values=conc_in_a, pores=net_a.pores('flow_outlet'), mode='overwrite')
                                ad_a_even.set_outflow_BC(pores = net_a.pores('internal', mode = 'nor'), mode='remove')
                                ad_a_even.set_outflow_BC(pores=net_a.pores('flow_inlet'), mode='overwrite')
                            if Flow_field == 1:
                                ad_a_even.set_value_BC(pores = net_a.pores('internal', mode = 'nor'), mode='remove')
                                ad_a_even.set_value_BC(values=conc_in_a, pores=net_a.pores('flow_inlet'), mode='overwrite')
                                ad_a_even.set_outflow_BC(pores = net_a.pores('internal', mode = 'nor'), mode='remove')
                                ad_a_even.set_outflow_BC(pores=net_a.pores('flow_outlet'), mode='overwrite')
                            
                            # Update reaction term with new values for eta
                            anolyte['pore.rxn_A1'] = cf.bv_rate_constant_ad_a(eta_a, Ai_a, rp_a, j0, km_factor, anolyte['pore.km_exp'], param)       
                            
                            # Solve the advection-diffusion-reaction equation
                            if itr > 0:
                                conc_guess = ad_a_even['pore.concentration']
                                ad_a_even.run(x0 = conc_guess)
                            else:
                                ad_a_even.run()
                                
                            # Update the concentration
                            conc_new_a = ad_a_even['pore.concentration']
                            
                        else:
                            # Retrieve the StokesFlow data for an odd network
                            anolyte.update(sf_a_odd.soln)
                            
                            if Larachi_HC == 0:             # OpenPNM fluid field physics
                                cf_pres_fit.Update_phase_with_SF(sf_a_odd, anolyte, 1, 0)
                                # MT coeficients                            
                                cf_pres_fit.Mass_transfer_coefficient(net_a, anolyte, 1, Velocity_dependent_km, 0, param)     
                                                       
                            elif Larachi_HC == 1:           # Mechanical energy balance fluid field
                                cf_pres_fit.Update_phase_with_SF(sf_a_odd, anolyte, 1, 1) 
                                # MT coeficients                            
                                cf_pres_fit.Mass_transfer_coefficient(net_a, anolyte, 1, Velocity_dependent_km, 1, param)  
                                
                            # Regenerate the (ad_dif_conductance) models
                            anolyte.regenerate_models('throat.ad_dif_conductance')
                            
                            # Set the boundary conditions for odd networks (normal case)
                            ad_a_odd.set_value_BC(pores = net_a.pores('internal', mode = 'nor'), mode='remove')    
                            ad_a_odd.set_value_BC(values=conc_in_a, pores=net_a.pores('flow_inlet'), mode='overwrite')
                            ad_a_odd.set_outflow_BC(pores = net_a.pores('internal', mode = 'nor'), mode='remove')
                            ad_a_odd.set_outflow_BC(pores=net_a.pores('flow_outlet'), mode='overwrite')
                            
                            # Update reaction term with new values for eta
                            anolyte['pore.rxn_A1'] = cf.bv_rate_constant_ad_a(eta_a, Ai_a, rp_a, j0, km_factor, anolyte['pore.km_exp'], param)       
                            
                            # Solve the advection-diffusion-reaction equation
                            if itr > 0:
                                conc_guess = ad_a_odd['pore.concentration']
                                ad_a_odd.run(x0 = conc_guess)
                            else:
                                ad_a_odd.run()
                            
                            # Update the concentration
                            conc_new_a = ad_a_odd['pore.concentration']
                        
                        # Update the value for the concentration
                        conc_a = conc_new_a * omega + conc_a * (1 - omega)
                        anolyte['pore.concentration'] = conc_a
                        
                        # Compute current according to the reaction of species
                        current_estimation_ad_a = cf.bv_rate_constant_oc_a(conc_a, eta_a, Ai_a, rp_a, j0, km_factor, anolyte['pore.km_exp'], param)[net_a.pores('internal')].sum()
                                    
                        # Compute oc reaction term with new values for conc_c
                        # The current generation is a source instead of a sink term, so minus bv_rate_constant_oc_anode
                        drdv = cf.bv_rate_derivative_oc_a(conc_a, eta_a, Ai_a, rp_a, j0, km_factor, anolyte['pore.km_exp'], param) 
                        r = -cf.bv_rate_constant_oc_a(conc_a, eta_a, Ai_a, rp_a, j0, km_factor, anolyte['pore.km_exp'], param)
                        
                        anolyte['pore.current_anolyte_A1'] = drdv
                        anolyte['pore.current_anolyte_A2'] = r - drdv * V_a
            
                        # Couple the potential field in both half cells via the membrane pores
                        V_membrane_pores_c = catholyte['pore.voltage'][net_c.pores('membrane')]
                        V_membrane_pores_a = V_membrane_pores_c - res_mem * Rm_factor * current_estimation_ad_a  
                        
                        # Update the membrane boundary condition for the anode
                        oc_a.set_value_BC(values=V_membrane_pores_a, pores=net_a.pores('membrane'), mode='overwrite')
                        
                        # Solve the ohmic conduction equation (potential field) for the liquid electrolyte in the porous cathode
                        oc_a.run()                        
                        
                        # Update voltage
                        V_new_a = oc_a['pore.voltage']
                        V_a = V_new_a * omega + V_a * (1 - omega)
                        anolyte['pore.voltage'] = V_a
            
                        # Compute new anodic overpotentials
                        eta_a = 0 - V_a
                        
                        # Compute current according to the transport of species
                        current_estimation_oc_a =  cf.bv_rate_constant_oc_a(conc_a, eta_a, Ai_a, rp_a, j0, km_factor, anolyte['pore.km_exp'], param)[net_a.pores('internal')].sum()
    
                        """-----------Convergence and updating--------------"""
                        # Calculate the relative error with the previous solution
                        rel_err_current_ad_a = cf.rel_error(current_estimation_ad_a, current_ad_a)
                        rel_err_current_oc_a = cf.rel_error(current_estimation_oc_a, current_oc_a)
                        rel_err_current_ad_c = cf.rel_error(current_estimation_ad_c, current_ad_c)
                        rel_err_current_oc_c = cf.rel_error(current_estimation_oc_c, current_oc_c)
                        
                        # Store the previous solution (iteration n)
                        current_ad_a = current_estimation_ad_a
                        current_oc_a = current_estimation_oc_a
                        current_ad_c = current_estimation_ad_c
                        current_oc_c = current_estimation_oc_c                        
                        
                        # Calculate the absolute error between the current density found in the
                        # anodic and cathodic compartment [A cm-2] (1e4 = conversion m2 to cm2)
                        abs_err_current_cat_an = abs((current_oc_a / A_ext_a / 1e4 - current_oc_c / A_ext_c / 1e4))                        
                        
                        # Check if found progression of the solution is within tolerances
                        convergence_ad_a = rel_err_current_ad_a < rel_tol
                        convergence_oc_a = rel_err_current_oc_a < rel_tol
                        convergence_ad_c = rel_err_current_ad_c < rel_tol
                        convergence_oc_c = rel_err_current_oc_c < rel_tol
                        convergence_a_c = abs_err_current_cat_an < abs_tol                        
                        
                        # Check for pores with a concentration greater than the imposed 
                        # inlet concentration. 
                        No_pores_cat_exceeding_inlet_conc = len(np.where(catholyte['pore.concentration']> param['conc_in_c'])[0])
                        No_pores_ano_exceeding_inlet_conc = len(np.where(anolyte['pore.concentration']> param['conc_in_a'])[0])
                        max_dev = 2.5e-2    # Maximum deviation of pore concentration from imposed inlet concentration
                        if No_pores_cat_exceeding_inlet_conc != 0:
                            max_conc_cat = catholyte['pore.concentration'][np.where(catholyte['pore.concentration']> param['conc_in_c'])].max()
                            Concentration_convergence_cat = (max_conc_cat - max_dev) < param['conc_in_c']
                        else:
                            Concentration_convergence_cat = True
                        
                        if No_pores_ano_exceeding_inlet_conc != 0:
                            max_conc_ano = anolyte['pore.concentration'][np.where(anolyte['pore.concentration']> param['conc_in_a'])].max()
                            Concentration_convergence_ano = (max_conc_ano - max_dev) < param['conc_in_a']
                        else:
                            Concentration_convergence_ano = True
                                                                                                                                                                                   # .......
                        if Concentration_convergence_ano and Concentration_convergence_cat and convergence_ad_a and convergence_oc_a and convergence_ad_c and convergence_oc_c and convergence_a_c and itr > Min_it_mass_charge or abs((abs(current_estimation_ad_c)-abs(current_estimation_ad_a)))/1e4/A_ext_c < 1e-5 and itr > Min_it_mass_charge and Concentration_convergence_ano and Concentration_convergence_cat:                    
                            print(f"CONVERGED network {network+1} at {E_cell} [V]  with {current_estimation_ad_c/A_ext_c/1e4:.5f} [A/cm2] and {itr} iterations")
                    
                            # Overpotential contributions:
                            # 1. Activation overpotential. It is found by minimzing the difference between 
                            # the ideal current (BV without MT effects) and the real current (BV with MT limitations)
                            # by fitting an activation overpotential for each pore (in the ideal current).
                            eta_act_c = np.zeros(net_c.Np)              # Initializng vector for saving activation overpotential
                            eta_act_a = np.zeros(net_a.Np)
                            
                            # The real current (BV with MT limitations)
                            anolyte['pore.current'] = cf.bv_rate_constant_oc_a(conc_a, eta_a, Ai_a, rp_a, j0, km_factor, anolyte['pore.km_exp'], param)
                            catholyte['pore.current'] = cf.bv_rate_constant_oc_c(conc_c, eta_c, Ai_c, rp_c, j0, km_factor, catholyte['pore.km_exp'], param)
                        
                            for pore in range(net_c.Np):
                                eta_act_c[pore] = optimize.fsolve(func = cf.find_eta_act, x0=eta_c[pore], args=('cathode', catholyte['pore.current'][pore], Ai_c[pore], j0, param))
                
                            for pore in range(net_a.Np):
                                eta_act_a[pore] = optimize.fsolve(func = cf.find_eta_act, x0=eta_a[pore], args=('anode', anolyte['pore.current'][pore], Ai_a[pore], j0, param))
                            
                            catholyte['pore.activation_overpotential'] = eta_act_c
                            anolyte['pore.activation_overpotential'] = eta_act_a                            
                            
                            # 2. Ohmic overpotential. It takes into account the electrolytic and membrane resistivity.
                            # The electrolytic resistivity is obtained by subtracting the pores electrolyte potential (V_c) 
                            # from the average potential of the pores neigboring the membrane.
                            # NOTE: In a symmetric cell setup, the applied voltage is equal to the total overpotential within the cell.
                            # Ohmic overpotential contribution:
                            catholyte['pore.ohmic_overpotential'] = V_c - V_membrane_pores_c.mean()
                            anolyte['pore.ohmic_overpotential'] = V_a  - V_membrane_pores_a.mean()
                            eta_ohm_mem = V_membrane_pores_c.mean()-V_membrane_pores_a.mean()                            
                            
                            # 3. Concentration overpotential. It is found by subtracting the overpotential found in
                            # the simulation from that of the activation overpotential:
                            catholyte['pore.concentration_overpotential'] = eta_c-catholyte['pore.activation_overpotential'] 
                            anolyte['pore.concentration_overpotential'] = eta_a-anolyte['pore.activation_overpotential']                                
                            
                            # Compute total current
                            # NOTE: Boundary pores are excluded, since these are artificial pores
                            # added by the SNOW algorithm, that are not part of the real network
                            current_sum_a = anolyte['pore.current'][net_a.pores('internal')].sum()
                            current_sum_c = catholyte['pore.current'][net_c.pores('internal')].sum()
                            current_sum = (abs(current_sum_a)+abs(current_sum_c))/2                            
                            current_sum_vec_int[:,network] = current_sum/A_ext_a/1e4
                            
                            # Output to Excel workbook
                            polarizationCurveData['current density'] = current_sum/A_ext_a/1e4
                            polarizationCurveData['cell voltage'] = abs(E_cell)
                            polarizationCurveData['anodic activation overpotential'] = anolyte['pore.activation_overpotential'][net_a.pores('internal')].mean() 
                            polarizationCurveData['cathodic activation overpotential'] = catholyte['pore.activation_overpotential'][net_c.pores('internal')].mean()
                            polarizationCurveData['anodic concentration overpotential'] = anolyte['pore.concentration_overpotential'][net_a.pores('internal')].mean()
                            polarizationCurveData['cathodic concentration overpotential'] = catholyte['pore.concentration_overpotential'][net_c.pores('internal')].mean()
                            polarizationCurveData['anodic ohmic overpotential'] = anolyte['pore.ohmic_overpotential'][net_a.pores('internal')].mean()
                            polarizationCurveData['cathodic ohmic overpotential'] = catholyte['pore.ohmic_overpotential'][net_c.pores('internal')].mean()
                            polarizationCurveData['membrane ohmic overpotential'] = abs(eta_ohm_mem)                              
                            
                            for idx, key in enumerate(polarizationCurveData):
                                ws.cell(row=1, column=idx+1).value = key
                                ws.cell(row=4+network+number_of_networks*E_cell_idx, column=idx+1).value = polarizationCurveData[key]
                            
                            # Update initial guesses for next cell voltage
                            eta0_a[:, network] = eta_a
                            eta0_c[:, network] = eta_c
                            conc0_a[:, network] = conc_a
                            conc0_c[:, network] = conc_c
            
                            if network is not number_of_networks-1:
                                if Flow_field == 0:
                                    if (network+1) % 2 == 0:
                                        conc0_in_a_odd[:, network+1] = conc_a[net_a.pores('flow_inlet')]
                                        conc0_in_c_odd[:, network+1] = conc_c[net_c.pores('flow_inlet')]
                                    else:
                                        conc0_in_a_even[:, network+1] = conc_a[net_a.pores('flow_outlet')]
                                        conc0_in_c_even[:, network+1] = conc_c[net_c.pores('flow_outlet')]
                                if Flow_field == 1:
                                    if (network+1) % 2 == 0:
                                        conc0_in_a_odd[:, network+1] = conc_in_a0         
                                        conc0_in_c_odd[:, network+1] = conc_in_c0         
                                    else:
                                        conc0_in_a_even[:, network+1] = conc_in_a0     
                                        conc0_in_c_even[:, network+1] = conc_in_c0       
            
                            # Outputting .VTK file for an applied voltage of -1 [V]
                            if (E_cell_idx+1)%10 == 0:
                                
                                ########### Cathodic network ##########
                                # Create to be outputted network and project:
                                pn_cat_vtk = op.network.Network()
                                pn_cat_vtk.update(net_c)
                                catholyte_vtk = op.phase.Phase(network=pn_cat_vtk, name='catholyte_vtk')
                                catholyte_vtk.update(catholyte)
    
                                ########### Anodic network ##########
                                # Create to be outputted network and project:
                                pn_ano_vtk = op.network.Network()
                                pn_ano_vtk.update(net_a)
                                anolyte_vtk = op.phase.Phase(network=pn_ano_vtk, name='anolyte_vtk')
                                anolyte_vtk.update(anolyte)      
                                
                                op.io.project_to_vtk(project = pn_cat_vtk.project, filename = '.\\output\\' + Output_folder + '\\' + name + '\\cathode_' + str(E_cell)[0:2] + '_' + str(E_cell)[3:] + 'V_' + 'network' + str(network) )
                                op.io.project_to_vtk(project = pn_ano_vtk.project, filename = '.\\output\\' + Output_folder + '\\' + name + '\\anode_' + str(E_cell)[0:2] + '_' + str(E_cell)[3:] + 'V'+ 'network' + str(network))
                                op.io.project_to_csv(project = pn_cat_vtk.project, filename = '.\\output\\' + Output_folder + '\\' + name + '\\cathode_' + str(E_cell)[0:2] + '_' + str(E_cell)[3:] + 'V_' + 'network' + str(network) )
                                op.io.project_to_csv(project = pn_ano_vtk.project, filename = '.\\output\\' + Output_folder + '\\' + name + '\\anode_' + str(E_cell)[0:2] + '_' + str(E_cell)[3:] + 'V'+ 'network' + str(network))
                            
                            break
                    current_sum_vec[:,E_cell_idx] = (current_sum_vec_int).mean()  
                        
            model = current_sum_vec           
            return model - data         
        
        import pandas as pd
        import lmfit
        
        # Import experimental current density vector [A/cm2]
        # Make sure that the voltage step between current_vec_exp and current_vec_model is the same 
        df = pd.read_excel('.//input//Experiment_fitting.xlsx')

        current_vec_exp = df['j'].to_numpy() 
        E_cell_vec_exp = - df['e'].to_numpy() 
        data = current_vec_exp
        method = 'leastsq'        
          
        params = lmfit.Parameters()
        # params.add('kmfactor', 1.0, min=0, max=1.5)
        params.add('Conductivityfactor', 1.10 , min=.9, max=1.50)
        
        out = lmfit.minimize(iterative_solver, params, method=method, args=(anolyte, catholyte, net_c, net_a, E_cell_vec_exp, rp_a, rp_c, j0, data, oc_c, oc_a,
                                                                            ad_c_even, ad_c_odd, ad_a_even, ad_a_odd, sf_c_odd, sf_c_even, sf_a_odd, sf_a_even, param, Velocity_dependent_km, Min_it_mass_charge, Larachi_HC))
        lmfit.report_fit(out)    
        final_time = time.time()
        print(f'Network simulations finished in {np.floor((final_time-starting_time)/60):.0f} minutes and {(final_time-starting_time)%60:.1f} seconds.')