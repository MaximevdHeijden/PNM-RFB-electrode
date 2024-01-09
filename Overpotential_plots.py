# Import functions
import openpnm as op
import numpy as np
from matplotlib import rc
from matplotlib import pyplot as plt
from matplotlib import cm as cm
from matplotlib import colors
import openpyxl
import pandas as pd
import pdb                                                  
import Custom_network_plotting as cn

# Load in network and phase via csv file. Note that only the network can be loaded
# in with .csv files. We thus copy the phase properties to the network, and all
# will be saved under project_c.

# Input of data
output_name = 'File_name_plot'
path = './input_plots/'
name_c = 'cathode_-1_0V_network0'
name_a = 'cathode_-1_0V_network0'

project_c = op.io.network_from_csv(filename = path + name_c + '.csv')
project_a = op.io.network_from_csv(filename = path + name_a + '.csv')

# Creating a new phase and network:
net_c = cn.Re_assign_network_V2_IDFF(project_c)
net_a = cn.Re_assign_network_V2_IDFF(project_a)
catholyte = cn.Re_assign_phase_V2(project_c, net_c, 'catholyte')
anolyte = cn.Re_assign_phase_V2(project_a, net_a, 'anolyte')

# General plot settings:
font_dict = {'sans-serif': 'Arial',
            'size': '12', 'weight': 'normal'}
rc('font', **font_dict) # Controls default text sizes
rc('axes', labelsize=18, linewidth=2, labelweight='normal')
rc('xtick', labelsize=18)
rc('xtick.major', width=2)
rc('ytick', labelsize=18)
rc('ytick.major', width=2)
rc('legend', fontsize=18)
rc('savefig', dpi=600)

# For the IDFF configuration we use the following axes. The current density will
# plotted over the width (the dimension (x) corresponding to the sides of the electrode)
def find_coords(net): 
    x = net['pore.coords'][net.pores('internal'), 1]
    y = net['pore.coords'][net.pores('internal'), 2]
    
    # correct network coordinates
    y = y+abs(y.min())
    z = net['pore.coords'][net.pores('internal'), 0]
    return x, y, z

# Get coordinates of anode and cathode
x_c, y_c, z_c = find_coords(net=net_c)
x_a, y_a, z_a = find_coords(net=net_a)

# correct network coordinates.
z_c = z_c - z_c.min(); 
z_a = z_a - z_a.min(); 

# Find width, length and thickness of anode and cathode
w_c = x_c.max() - x_c.min()
l_c = y_c.max() - y_c.min() 
t_c = z_c.max() - z_c.min()
w_a = x_a.max() - x_a.min()
l_a = y_a.max() - y_a.min()
t_a = z_a.max() - z_a.min() 

max_length = min(l_c, l_a)
# Define the performance indicators
# Overpotential fields [V]
act_eta_c = (catholyte['pore.activation_overpotential'][net_c.pores('internal')])
ohm_eta_c = (project_c['pore.ohmic_overpotential'][net_c.pores('internal')])
conc_eta_c = (catholyte['pore.concentration_overpotential'][net_c.pores('internal')])
pore_voltage = catholyte['pore.voltage'][net_c.pores('internal')]
test =  ohm_eta_c + act_eta_c + conc_eta_c

# Convert units
gridsize = 16

# Current distribition in 1D:
def current_distribution(bin_width, z, act, ohm, conc, tot, som, pores):     
    
    # Set up iterating variables
    max_z = int(max(np.ceil(z/bin_width)*bin_width))
    z_range = range(0, max_z, bin_width) 
    act_1 = np.zeros(len(z_range)) 
    ohm_1 = np.zeros(len(z_range)) 
    conc_1 = np.zeros(len(z_range)) 
    tot_1 = np.zeros(len(z_range)) 
    som_1 = np.zeros(len(z_range)) 
    bins = np.zeros(len(z_range))
    j = 0
    a = 1
    # Divide pores over bins
    for size_bin in z_range:
        for pore in pores:
            if size_bin < z[pore] <= size_bin + bin_width:
                bins[j] += 1
                act_1[j] += act[pore]
                ohm_1[j] += ohm[pore]
                conc_1[j] += conc[pore]
                tot_1[j] += tot[pore]
                som_1[j] += som[pore]                
                a = a+1
            act_1[j] = act_1[j]/a
            ohm_1[j] = ohm_1[j]/a
            conc_1[j] = conc_1[j]/a
            tot_1[j] = tot_1[j]/a
            som_1[j] = som_1[j]/a           
            a=1    
        j += 1
        
    cd = {}
    cd['z_range'] = z_range
    cd['bins'] = bins
    cd['act'] = act_1
    cd['ohm'] = ohm_1
    cd['conc'] = conc_1
    cd['tot'] = tot_1
    cd['som'] = som_1
    return cd

# Current distribution
z_c_final = z_c*1e6
cd = current_distribution(bin_width = 1, z=z_c_final, act=act_eta_c, ohm=ohm_eta_c, conc=conc_eta_c, tot=pore_voltage, som=test, pores=net_c.pores('internal'))

# Plot length-thickness hexagonal binning plots for the current field
fig, ax = plt.subplots(figsize=[6,4])
ax.plot(cd['z_range'], cd['act'] , 'r', linewidth=4)
ax.plot(cd['z_range'], cd['ohm'] , 'g', linewidth=4)
ax.plot(cd['z_range'], cd['conc'] , 'b', linewidth=4)
ax.plot(cd['z_range'], cd['tot'] , 'k', linewidth=4)
# ax.plot(cd['z_range'], cd['som'] , 'k', linewidth=4)
ax.set_xlabel('Electrode thickness [um]')
ax.set_ylabel('Overpotentials [V]')
# ax.xaxis.set_ticks([0, 200])
# ax.yaxis.set_ticks([0, 0.1, 0.2,0.3])
# ax.yaxis.set_label_coords(-0.2, 0.2)
plt.savefig(fname='.\\plots\\' + output_name + '.tif', bbox_inches='tight')
plt.show()
    
wb = openpyxl.Workbook()
ws = wb.active
polarizationCurveData = {}

# Output to Excel workbook                        
for idx in range(len(cd['z_range'])):
    ws.cell(row=idx+2, column=1).value = cd['z_range'][idx]
    ws.cell(row=idx+2, column=2).value = cd['act'][idx]
    ws.cell(row=idx+2, column=3).value = cd['conc'][idx]

wb.save(filename = '.\\plots\\' + output_name + 'act_mt_cathode.xlsx')