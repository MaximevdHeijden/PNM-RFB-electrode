"""
Notebook serves to obtain the following network properties from the extracted 
network( which performed in @Network_Extraction.py):
    1. Connectivity
    2. Pore size distribution
    3. Porosity
    4. Permeability 
Further, it converts the .csv file to a .pnm (v3.0.0) and .vtk file, with the same 
input name as the .csv file.

Notebook is divided into the following steps:
    1. Loading in .csv and .tif network files and creating .xlsx (Excel)
       directory for storing the properties
    2. Outputting .pnm network
    3. Defining functions for obtaining the properties
    4. Running functions to extract the properties
    5. Outputting extracted data to .xlsx file
    6. Outputting .vtk file of network 
    
OpenPNM v3.0.0 cannot load projects (.pnm) from OpenPNM v2.2.0. Hence, networks  
will be imported via .CSV files. To enable this import add the following to 
op.io.network_from_csv: in a = read_table add as argument: low_memory=False. 
Further, labels and properties will be mixed in the extracted network, so a new
(copy) of the network is created in the module costum_network. 

Input: 
    1. .csv file of the extracted network
    2. .tif file of the to be extracted network
    
Output: An Excel file containing the above named properties. 

Options:
    The 17x_vertical_Freudenberg contains some unreasonably large throats (2) which need to be trimmed
    ToBeTrimmed = 0 ->   You do NOT import the 17x_vertical_Freudenberg.csv file and require no trimming
    ToBeTrimmed = 1 ->   You DO import the 17x_vertical_Freudenberg.csv file and require trimming
"""

import openpnm as op
import porespy as ps
import numpy as np
from skimage import io
from skimage import util
import openpyxl
import matplotlib.pyplot as plt
import os
import Costum_network as cn
import Costum_functions as cf

############################# Network alterations functions ###################
ToBeTrimmed = 0                             # Trim unrealstic larges pores. 
# If ToBeTrimmed == 0 -> No trimming takes place
# If ToBeTrimmed == 1 -> Trimming takes place, where an indicated number of largest pores are trimmed from the network.

Shift = 0                                   # Perform a 3D rotation on the extracted network. 
# The interdigitated flow field may require rotationd depending on the stitching of network in ImageJ 
# If Shift == 0 -> No rotation takes place
# If Shift == 1 -> Rotation takes place over an indicated angle over the x-axes. 

"""-------------------Loading in network and image--------------------------"""
cwd = os.getcwd()                           # Curent Working directory
network_path = '\\input\\'                 
file_name ='Freudenberg'                    # CSV file from V2.2.0 (but altered):
directory = cwd + network_path + file_name

# The extracted network from .CSV contains mixed properties and labels. A new
# network is created with correctly assigned properties and labels.
net_csv = op.io.network_from_csv(filename = directory)
net = cn.Re_assign_network_V2(net_csv)
net.regenerate_models()

# Load in image
network_path_im = '\\input\\'               # Folder with stored file under cwd
file_name_im = 'Freudenberg.tif'            # Name of image
directory_im = cwd + network_path_im + file_name_im
im_solid = io.imread(fname = directory_im)  # Reading in image 
voxel_size = 1.65e-6                        # Voxel size in [m] 

# Name output .pnm and .vtp file
network_path_network = '\\output_version_3_0\\'
directory_network = cwd + network_path_network + file_name

# Name of the to be outputed Excel file             
directory_excel = cwd +  network_path_network + file_name

if im_solid.dtype is not bool:
    print("Converting supplied image to boolean")
    im_solid = im_solid > 0                 # Convert to Boolean datatype
# Invert so that the void space is set as True'
im_void = util.invert(im_solid)

# Trimming the unrealistically large pores in the 17x_vertical_Freudenberg
if ToBeTrimmed == 1:
    # Removing the two extremely long throats (and its boundary pore, to avoid singularity)
    # Probably occurance hereof is an error in extraction.
    for i in range(2):
        index_throat_max = np.argwhere(net['throat.length'] == net['throat.length'].max())
        Pores = net.find_connected_pores(throats= index_throat_max)
        op.topotools.trim(network=net, pores = Pores[0][1])

if (Shift == 1): 
    # Function rotates the network angle around x-axes and re-assigns 
    # the correct labels from a positive z-direction reference.  
    # Thereafter the network is shifted such that its boundary plane coincides with the origin. 
    # The shift value is obtained through trial and error, and checking the minimal x-location of the network (which should be 0)
    net = cn.Rotating_network(net, np.pi/2, 1, 0.00101145)
    
# Checking network health:
health_net = op.utils.check_network_health(network = net)
print(health_net)

"""------------------------Outputting .pnm netwok---------------------------"""
op.Workspace().save_project(project = net.project, 
                            filename = directory_network)

"""-----------------------Function definitions------------------------------"""
def find_connectivity(net):
    r''' find_connectivity finds the connectivity of every pore in a network.

    Parameters
    ----------
    net : The network for which you would like to find the connectivity.

    Returns
    -------
    unique : Amount of pores a pore is connected to.
    counts : The amount of pores which have this unique connectivity.
    (e.g. for unique = 1, counts = 3, 3 pores have a connectivity of 1)
    norm_counts : counts normalized to unity. 
    '''
    
    # Initializing the connectivity matrix and filling it
    conn = np.zeros(len(net.pores('internal'))) 
    conn = net.num_neighbors(net.pores('internal'))
    
    # Finding all unique connectivities and the no. of counts of each connectivity
    unique, counts = np.unique(conn, return_counts=True)
    
    norm_counts = counts/np.sum(counts)
    # norm_counts_times_conn = counts*unique/sum(counts*unique)
    
    return unique, norm_counts, conn


def pore_size_distribution(bin_width, pore_volume, pore_diameter, pores):     
    r''' pore_size_distribution returns the normalized volume and cumulative
    normalized volume for a given set of pores, which can be used to plot the
    pore size distribution histogram.
    Parameters
    ----------
    bin_width : Specified bin_width for the histogram.
    pores : The pores that are analyzed.
    pore_volume : Array of pore volumes of every pore in pores.
    pore_diameter : Array of pore diameters of every pore in pores.
            
    Returns
    -------
    pore_size_range : The position of the left side of the bins.
    bins : The number of pores in every bin 
    norm_volume : The normalized volume that every bin accounts for.
    cum_volume : THe cumulative normalized volume that every bin up to that bin
    accounts for. 
    '''
    
    pore_diameter_2 = pore_diameter * 1e6               # Conversion from [m] to [um]
    norm_pore_volume = pore_volume/np.sum(pore_volume)  # Normalize pore volume to unity
    
    # Set up iterating variables
    max_pore_size = int(max(np.ceil(pore_diameter_2/bin_width)*bin_width))
    pore_size_range = range(0, max_pore_size, bin_width)
    norm_volume = np.zeros(len(pore_size_range)) 
    bins = np.zeros(len(pore_size_range))
    i = 0
    
    # Divide pores over bins
    for size_bin in pore_size_range:
        for pore in pores:
            if size_bin < pore_diameter_2[pore] <= size_bin + bin_width:
                bins[i] += 1
                norm_volume[i] += norm_pore_volume[pore]
        i += 1
        
    # Test if all pores are assigned    
    print(f'Total normalized volume (should be 1.00): {sum(norm_volume) :.2f}')
    
    # Compute cumulative volume
    j = 0
    cum_volume_iter = 0
    cum_volume = np.zeros(len(pore_size_range))
    
    for bin_volume in norm_volume:
        cum_volume[j] = cum_volume_iter + bin_volume
        cum_volume_iter += bin_volume
        j += 1
    
    psd = {}
    psd['pore_size_range'] = pore_size_range
    psd['bins'] = bins
    psd['norm_volume'] = norm_volume
    psd['cum_volume'] = cum_volume
    
    return psd

    
def find_porosity(net, im, voxel_size):
    r'''
    Finds the porosity calculated from the network and from the corresponding
    image.

    Parameters
    ----------
    net : Pore network of the electrode of interest.
    im : Corresponding image with the void space set as True.
    voxel_size : Voxel_size of the image in [meter].

    Returns
    -------
    net_porosity : Porosity calculated from the network.
    im_porisity : Porosity calculated from the image.
    '''
    
    pore_volume = net['pore.volume'][net.pores('internal')] 
    
    void_volume = pore_volume.sum()
    total_volume = im.shape[0]*im.shape[1]*im.shape[2]*voxel_size**3
    
    net_porosity = void_volume/total_volume
    im_porosity = ps.metrics.porosity(im)
    
    return net_porosity, im_porosity


def Darcy_solver(Inlet_pores, A, L, del_P, phase, alg):
    r'''
    Find the permeability from a StokesFlow algorithm in one (1) direction. Gives 
    the exact same permeability calculation as the calc_effective_permeability
    method which was still available in V2.2.0. 

    Parameters
    ----------
    Inlet_pores : 
        Pores where the highest pressure boundary condition is applied.
    A : 
        Area of the applied flow, obtained from the loaded image [m2].
    L : 
        Path length of the fluid flow [m].
    del_P : 
        The pressure gradient applied [Pa].
    phase : 
        The phase to which the results of the StokesFlow algorithm has been
        saved.
    alg:
        The StokesFlow algorithm.
   
    Returns
    -------
    K:
        The permeability of the flow direction.
    '''
    Q = alg.rate(pores=Inlet_pores,mode='group')[0]     # Flow rate @inlet pores [m3/s]
    mu_w = phase['throat.viscosity']                    # Viscosity in the throats [Pa.s]
    K = Q*mu_w*L/(A*del_P)
    
    return K
    

def permeability(net, im, voxel_size, phase):
    r'''
    Finds the permeability vector for the given network.
    
    NOTE: Check if net.pores('left, right, front, back, bottom and top') correspond
    with the right faces for inlet and outlet with the given image shape in Paraview
    before running! 
    
    In this case:   'left, right' = x direction
                    'front, back' = y direction
                    'top, bottom' = z direction
    
    Parameters
    ----------
    net : Pore network of the electrode of interest.
    im : Corresponding image with the void space set as True.
    voxel_size : Voxel_size of the image in [meter].
    phase: Phase acting on the pore network model
    Returns
    -------
    K: Permeability vector [Kx, Ky, Kz]
    '''
    
    # Network flow cross-sectional area
    area = {
    'x': im.shape[1] * im.shape[2] * voxel_size**2,
    'y': im.shape[0] * im.shape[2] * voxel_size**2,
    'z': im.shape[0] * im.shape[1] * voxel_size**2, 
        }

    # Network flow length
    length = {
    'x': im.shape[0] * voxel_size,
    'y': im.shape[1] * voxel_size,
    'z': im.shape[2] * voxel_size,
        }
    
    # Checking dimensions of the network:
    x_inlet = net.pores('left'); x_outlet = net.pores('right');
    y_inlet = net.pores('front'); y_outlet = net.pores('back');
    z_inlet = net.pores('top'); z_outlet = net.pores('bottom');    
    L_x = op.topotools.get_domain_length(net, inlets=x_inlet, outlets=x_outlet)
    L_y = op.topotools.get_domain_length(net, inlets=y_inlet, outlets=y_outlet)
    L_z = op.topotools.get_domain_length(net, inlets=z_inlet, outlets=z_outlet)
    L_net = {
    'x': L_x,
    'y': L_y,
    'z': L_z, 
        }
    
    print('NOTE: When running permeability(), make sure that the assigned pores'
          ' match the dimensions (image shape) given in this function!:',
          '\nImage:  ', length, '\nNetwork:',L_net)
        
    # Perform StokesFlow algorithm to calculate x Permeability
    flow_x = op.algorithms.StokesFlow(phase = phase, network = net)
    flow_x.set_value_BC(pores = x_inlet, values = 101325)
    flow_x.set_value_BC(pores = x_outlet, values = 0)
    flow_x.run()
    phase.update(flow_x.soln)
    K_x = Darcy_solver(Inlet_pores = x_inlet, A = area['x'] , L = length['x'],
                       del_P = 101325, phase = phase, alg = flow_x)[0]
    #K_x = flow_x.calc_effective_permeability(domain_area=area['x'], domain_length=length['x'])[0]

    # Perform StokesFlow algorithm to calculate y Permeability
    flow_y = op.algorithms.StokesFlow(phase = phase, network = net)
    flow_y.set_value_BC(pores = y_inlet, values = 101325)
    flow_y.set_value_BC(pores = y_outlet, values = 0)
    flow_y.run()
    phase.update(flow_y.soln)
    K_y = Darcy_solver(Inlet_pores = y_inlet, A = area['y'] , L = length['y'],
                       del_P = 101325, phase = phase, alg = flow_y)[0]
    
    # Perform StokesFlow algorithm to calculate z Permeability
    flow_z = op.algorithms.StokesFlow(phase = phase, network = net)
    flow_z.set_value_BC(pores = z_inlet, values = 101325)
    flow_z.set_value_BC(pores = z_outlet, values = 0)
    flow_z.run()
    phase.update(flow_z.soln)
    K_z = Darcy_solver(Inlet_pores = z_inlet, A = area['z'] , L = length['z'],
                       del_P = 101325, phase = phase, alg = flow_z)[0]    
    K = [K_x*1e12, K_y*1e12, K_z*1e12]
    
    net.regenerate_models() 
    air.regenerate_models()
    
    return K 


"""-----------------------Adding physics and phase--------------------------"""
# The following models and phase objects will be added to the simulation:
#   Air:
#       1. Adding "Air" phase with an assigned viscosity
#       2. Add Air property models
#
#   Physics:
#       1. Flow shape factors
#       2. Hydraulic conductance
#       3. Poisson shape factors
#       4. Diffusive conductance
#       5. Advection diffusion
#       6. Electrical conductance

# Creating air phase
air = op.phase.Air(network = net)
air['pore.surface_tension'] = 0.072
air.interpolate_data('throat.surface_tension')
air['pore.molecular_weight'] = 0.0291
air['pore.diffusivity'] = 2.06754784e-05
air['throat.diffusivity'] = 2.06754784e-05
air['pore.electrical_conductivity'] = 28        # S/m
#air.regenerate_models()

# Physics
f_hyd = cf.Flow_shape_factors_ball_and_stick
air.add_model(propname = 'throat.flow_shape_factors', model = f_hyd)    
f_Hyd_cond = cf.Hydraulic_conductance_Hagen_Poiseuille
air.add_model(propname = 'throat.hydraulic_conductance', model = f_Hyd_cond)
f_poisson = cf.Poisson_shape_factors_ball_and_stick
air.add_model(propname = 'throat.poisson_shape_factors', model = f_poisson)
f_diff_cond = cf.Diffusive_conductance_mixed_diffusion
air.add_model(propname = 'throat.diffusive_conductance', model = f_diff_cond)
f_ad_dif = cf.Advection_diffusion
air.add_model(propname = 'throat.ad_dif_conductance', model = f_ad_dif)
f_elec_con = cf.Electrical_conductance_series_resistors
air.add_model(propname = 'throat.electrical_conductance', model = f_elec_con)

net.regenerate_models() 
air.regenerate_models()

op.models.collections.physics.standard
# f_hyd = op.models.geometry.hydraulic_size_factors.cones_and_cylinders      #spheres_and_cylinders
# net.add_model(propname = 'throat.hydraulic_size_factors', model = f_hyd)    
# f_dif_con = op.models.geometry.diffusive_size_factors.cones_and_cylinders  #spheres_and_cylinders
# net.add_model(propname = 'throat.diffusive_size_factors', model = f_dif_con)
# net.regenerate_models() #This net() fucks up the calculation -> it 
# # enters Nans in the size factors.

"""------------------------Extracting properties----------------------------"""
# 1. Connectivity
# First a label is created for the internal pores: selecting all non-boundary pores. 
conn_unique, conn_counts, conn = find_connectivity(net=net)
avg_connectivity = np.mean(conn)
print(f'Average connectivity: {avg_connectivity :.2f}')

# 2. Pore size distribution
psd = pore_size_distribution(bin_width = 2, 
                              pore_volume=net['pore.volume'][net.pores('internal')], 
                              pore_diameter=net['pore.diameter'][net.pores('internal')],
                              pores=net.pores('internal'))
avg_pore_size = np.mean(net['pore.diameter'][net.pores('internal')])
print(f'Average pore size: {avg_pore_size*1e6 :.2f} um')

# 3.Porosity 
net_porosity, im_porosity = find_porosity(net=net, im=im_void, voxel_size=voxel_size)
x_length = np.linspace(0, im_void.shape[0]*voxel_size*1e6, im_void.shape[0]) # Length of network [um]
y_length = np.linspace(0, im_void.shape[1]*voxel_size*1e6, im_void.shape[1])
z_length = np.linspace(0, im_void.shape[2]*voxel_size*1e6, im_void.shape[2])
x_profile = ps.metrics.porosity_profile(im_void, 0)
y_profile = ps.metrics.porosity_profile(im_void, 1)
z_profile = ps.metrics.porosity_profile(im_void, 2)
print(f'Image Porosity: {im_porosity*100. :.2f} % \nNetwork Porosity: {net_porosity*100 :.2f} %')

# 4. Internal surface area
surf_area = net['pore.surface_area'].sum()
ECSA = surf_area/(im_void.shape[0] * im_void.shape[1] * im_void.shape[2] * voxel_size**3) # [m2/m3]
print(f'Internal surface area: {ECSA*0.01 :.2f} cm2/cm3')

# 5. Permeability
K = permeability(net=net, im=im_void, voxel_size=voxel_size, phase = air)
formatted_K = [ '%.2f' % elem for elem in K ]
print(f'Permeability vector: {formatted_K} Darcy')

"""------------------------Outputting properties----------------------------"""
# All properties will be outputted to a .xls file
wb = openpyxl.Workbook()
ws = wb.active

output_variables = [conn_unique, conn_counts,  psd['pore_size_range'], 
                    psd['norm_volume'], psd['cum_volume'], x_length, 
                    x_profile*100, y_length, y_profile*100, 
                    z_length, z_profile*100, K]

output_names = ['pore connections', 'frequency', 'pore size (PSD)', 
                'normalized volume', 'cumulative volume', 'x length (porosity)',
                'x profile', 'y length', 'yprofile',
                'z length', 'z profile', 'permeability [x,y,z]']

units = ['-', '-', 'um',
          '-', '-', 'um',
          '%', 'um', '%',
          'um', '%', 'Da']

ws.cell(row=1,column=1).value = 'average connectivity (-)'
ws.cell(row=1,column=2).value = avg_connectivity
ws.cell(row=2,column=1).value = 'average pore size (um)'
ws.cell(row=2,column=2).value = avg_pore_size
ws.cell(row=3,column=1).value = 'network porosity (%)'
ws.cell(row=3,column=2).value = net_porosity*100
ws.cell(row=4,column=1).value = 'Internal surface area (cm2/cm3)'
ws.cell(row=4,column=2).value = ECSA*0.01
ws.cell(row=5,column=1).value = 'image porosity (%)'
ws.cell(row=5,column=2).value = im_porosity*100

for column_num in range(0, len(output_variables)): # column_num represents the variables in ouput_variables.
    ws.cell(row=1, column=column_num+4).value = output_names[column_num]
    ws.cell(row=2, column=column_num+4).value = units[column_num]
    ws.cell(row=3, column=column_num+4).value = None
    ws.cell(row=4, column=column_num+4).value = None
    for row_num in range(0, len(output_variables[column_num])): # row_num represents the ith entry within the variable array
        ws.cell(row = row_num + 4, column = column_num + 4).value = output_variables[column_num][row_num] # row_num + 5 to convert to Origin format 

wb.save(filename = directory_excel + '_network_properties.xlsx')
print(f'information saved to .\\output\\{file_name}_network_properties.xlsx')

"""------------------------Outputting .vtk netwok---------------------------"""
####################Saving the file as a .vtk project#########################
# Overlaying the network and the image is requires using paraview since the image
# is in 3D. The is one gotcha due to the differences in how paraview and numpy
# number axes: it is necessary to rotate the image using 
# ps.tools.align_image_with_openpnm. Further, we can only export projects. This 
# requires the presence of a Phase. Here we will just add an empty generic phase
# object. 
op.io.project_to_vtk(project = net.project, filename = directory_network )